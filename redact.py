#!/usr/bin/env python3
"""
PDF Credit Card Bill Redaction Tool

Redacts sensitive transaction descriptions and cities from ICS credit card bills
based on configurable trigger terms. Uses true redaction that removes text from
the PDF content stream (not just overlay).

Usage:
    python redact.py bill.pdf
    python redact.py bill.pdf --config custom_terms.yaml
    python redact.py *.pdf --output redacted/
    python redact.py bill.pdf --dry-run
"""

import argparse
import re
import sys
from collections import defaultdict
from pathlib import Path

import fitz  # PyMuPDF
import yaml
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# Dutch month abbreviations used in ICS bills
# Note: "mei" (May) is the full month name and doesn't have a period
DATE_PATTERN = re.compile(
    r"^\d{2}\s+(jan|feb|mrt|apr|mei|jun|jul|aug|sep|okt|nov|dec)\.?$",
    re.IGNORECASE
)

# Pattern to detect exchange rate lines
EXCHANGE_RATE_PATTERN = re.compile(r"^Wisselkoers\s+[A-Z]{3}", re.IGNORECASE)

# Default config if no config file provided
DEFAULT_CONFIG = {
    "exceptions": [
        "ONLYFANS", "FANSLY", "TINDER", "BUMBLE", "GRINDR",
        "BETWAY", "POKERSTARS", "BET365", "CASINO", "LOTTERY",
        "PORNHUB", "BRAZZERS", "STRIP", "XXX",
        "PHARMACY", "CLINIC", "THERAPY", "MENTAL", "PSYCHIATR"
    ],
    "exact_match": ["OF"],
    "redaction": {
        "default": "none",
        "color": [0, 0, 0],
        "include_exchange_rate": True
    }
}


def parse_color(color_value) -> tuple:
    """Parse color from config (RGB list or hex string) to tuple."""
    if isinstance(color_value, list) and len(color_value) >= 3:
        return tuple(float(c) for c in color_value[:3])
    elif isinstance(color_value, str):
        # Parse hex color like "#FF0000" or "FF0000"
        hex_str = color_value.lstrip("#")
        if len(hex_str) == 6:
            r = int(hex_str[0:2], 16) / 255.0
            g = int(hex_str[2:4], 16) / 255.0
            b = int(hex_str[4:6], 16) / 255.0
            return (r, g, b)
    return (0, 0, 0)  # Default black


def get_redaction_settings(config: dict) -> dict:
    """Extract redaction settings from config with defaults."""
    redaction = config.get("redaction", {})
    return {
        "color": parse_color(redaction.get("color", [0, 0, 0])),
        "include_exchange_rate": redaction.get("include_exchange_rate", True)
    }


def load_config(config_path: Path | None) -> dict:
    """Load redaction configuration from YAML file or use defaults."""
    if config_path and config_path.exists():
        with open(config_path) as f:
            return yaml.safe_load(f)
    return DEFAULT_CONFIG


def extract_text_spans(page: fitz.Page) -> list[dict]:
    """Extract all text spans with their bounding boxes from a page."""
    spans = []
    blocks = page.get_text("dict", flags=fitz.TEXT_PRESERVE_WHITESPACE)["blocks"]

    for block in blocks:
        if block.get("type") != 0:  # Skip non-text blocks
            continue
        for line in block.get("lines", []):
            for span in line.get("spans", []):
                spans.append({
                    "text": span["text"].strip(),
                    "bbox": fitz.Rect(span["bbox"]),
                    "origin": span.get("origin", (span["bbox"][0], span["bbox"][1]))
                })
    return spans


def group_spans_by_row(spans: list[dict], tolerance: float = 3.0) -> dict[float, list[dict]]:
    """Group text spans by their Y-coordinate (row)."""
    rows = defaultdict(list)

    for span in spans:
        y = span["bbox"].y0
        # Find existing row within tolerance
        matched_row = None
        for existing_y in rows.keys():
            if abs(existing_y - y) <= tolerance:
                matched_row = existing_y
                break

        if matched_row is not None:
            rows[matched_row].append(span)
        else:
            rows[y].append(span)

    # Sort spans within each row by X coordinate
    for y in rows:
        rows[y].sort(key=lambda s: s["bbox"].x0)

    return dict(sorted(rows.items()))


def is_transaction_row(spans: list[dict]) -> bool:
    """Check if a row is a transaction row (starts with a date)."""
    if not spans:
        return False
    first_text = spans[0]["text"]
    return bool(DATE_PATTERN.match(first_text))


def parse_transaction_row(spans: list[dict]) -> dict | None:
    """
    Parse a transaction row and extract column information.
    Returns dict with dates, description, city, country, amounts, bounding boxes, and column boundaries.
    """
    if len(spans) < 5:
        return None

    # Find the first two date columns
    date_cols = []
    for i, span in enumerate(spans):
        if DATE_PATTERN.match(span["text"]):
            date_cols.append(i)
        if len(date_cols) >= 2:
            break

    if len(date_cols) < 2:
        return None

    # After the two date columns, we have: description, city, country, amounts...
    desc_start_idx = date_cols[1] + 1

    if desc_start_idx >= len(spans):
        return None

    # Find country code column (3-letter uppercase)
    country_idx = None
    for i in range(desc_start_idx, len(spans)):
        text = spans[i]["text"]
        if re.match(r"^[A-Z]{3}$", text):
            country_idx = i
            break

    if country_idx is None or country_idx <= desc_start_idx:
        return None

    # City is typically the span right before the country code
    # Description is everything between date2 and city
    city_idx = country_idx - 1

    # Collect description spans (may be multiple spans)
    desc_spans = spans[desc_start_idx:city_idx]
    city_span = spans[city_idx]
    country_span = spans[country_idx]
    date1_span = spans[date_cols[0]]
    date2_span = spans[date_cols[1]]

    if not desc_spans:
        return None

    # Combine description text
    description = " ".join(s["text"] for s in desc_spans)
    city = city_span["text"]
    country = country_span["text"]

    # Extract dates
    date1 = date1_span["text"]
    date2 = date2_span["text"]

    # Extract amounts (everything after country code)
    amount_spans = spans[country_idx + 1:]
    amounts = [s["text"] for s in amount_spans]

    # Calculate bounding boxes for description and city columns
    desc_bbox = fitz.Rect(desc_spans[0]["bbox"])
    for s in desc_spans[1:]:
        desc_bbox |= s["bbox"]  # Union of rectangles

    city_bbox = city_span["bbox"]

    # Calculate full cell boundaries:
    # - Description cell starts right after date2 column, ends at city cell start
    # - City cell starts at city span, ends at country column start
    row_height = max(s["bbox"].height for s in spans)
    row_y0 = min(s["bbox"].y0 for s in spans)
    row_y1 = row_y0 + row_height

    # Full description cell: from end of date2 to start of city
    desc_cell_x0 = date2_span["bbox"].x1 + 2  # Small gap after date
    desc_cell_x1 = city_bbox.x0 - 2  # Small gap before city

    # Full city cell: from city start to country start
    city_cell_x0 = city_bbox.x0 - 2
    city_cell_x1 = country_span["bbox"].x0 - 2

    return {
        "date1": date1,
        "date2": date2,
        "description": description,
        "city": city,
        "country": country,
        "amounts": amounts,
        "desc_bbox": desc_bbox,
        "city_bbox": city_bbox,
        "desc_spans": desc_spans,
        "city_span": city_span,
        "row_y0": row_y0,
        "row_y1": row_y1,
        # Full cell boundaries for complete redaction
        "desc_cell": fitz.Rect(desc_cell_x0, row_y0, desc_cell_x1, row_y1),
        "city_cell": fitz.Rect(city_cell_x0, row_y0, city_cell_x1, row_y1)
    }


def should_redact(description: str, city: str, config: dict) -> tuple[bool, str]:
    """
    Check if a transaction should be redacted based on config.

    default: "none" → redact nothing, exceptions specify what TO redact
    default: "all" → redact everything, exceptions specify what to KEEP

    Returns (should_redact, matched_term).
    """
    desc_upper = description.upper()
    city_upper = city.upper()

    default = config.get("redaction", {}).get("default", "none")

    # Get exceptions list (with backward compatibility for redact_terms)
    exceptions = config.get("exceptions") or config.get("redact_terms") or []
    exact_match = config.get("exact_match") or []

    # Check if entry matches any exception
    matched_term = ""

    # Check exact matches first
    for term in exact_match:
        if desc_upper == term.upper():
            matched_term = f"exact:{term}"
            break

    # Check substring matches
    if not matched_term:
        for term in exceptions:
            term_upper = term.upper()
            if term_upper in desc_upper or term_upper in city_upper:
                matched_term = term
                break

    if default == "all":
        # Redact everything EXCEPT exceptions
        if matched_term:
            return False, ""  # Exception found - DON'T redact
        return True, "[default: all]"  # No exception - redact
    else:
        # default: "none" - Redact nothing EXCEPT exceptions
        if matched_term:
            return True, matched_term  # Exception found - redact
        return False, ""  # No exception - don't redact


def is_exchange_rate_row(spans: list[dict]) -> bool:
    """Check if a row is an exchange rate line (Wisselkoers...)."""
    if not spans:
        return False
    first_text = spans[0]["text"]
    return bool(EXCHANGE_RATE_PATTERN.match(first_text))


def verify_redaction(output_path: Path, redacted_items: list[dict]) -> dict:
    """
    Verify that redacted content is truly removed from the saved PDF.
    Opens the saved file and searches for the exact redacted text strings.
    Only flags as leaked if the specific description text is found (cities may
    legitimately appear in other non-redacted transactions).
    Returns verification result with pass/fail status and any leaked terms found.
    """
    verification = {
        "passed": True,
        "checked_terms": [],
        "leaked_terms": []
    }

    try:
        doc = fitz.open(output_path)
        full_text = ""
        for page in doc:
            full_text += page.get_text()
        doc.close()

        # Convert to uppercase for case-insensitive comparison
        full_text_upper = full_text.upper()

        # Check each redacted item - focus on the unique description text
        # Cities like "LONDON" may appear in other non-redacted transactions
        for item in redacted_items:
            desc = item["description"]

            # Check if the full description string appears (most reliable check)
            # For very short descriptions (like "OF"), also check it as a word boundary
            if len(desc) >= 2:
                verification["checked_terms"].append(desc)
                # For short terms, check with word boundaries to avoid false matches
                if len(desc) <= 3:
                    # Check if it appears as a standalone word
                    import re
                    pattern = r'\b' + re.escape(desc.upper()) + r'\b'
                    if re.search(pattern, full_text_upper):
                        verification["passed"] = False
                        verification["leaked_terms"].append(f"desc:{desc}")
                elif desc.upper() in full_text_upper:
                    verification["passed"] = False
                    verification["leaked_terms"].append(f"desc:{desc}")

    except Exception as e:
        verification["passed"] = False
        verification["error"] = str(e)

    return verification


def get_exchange_rate_redact_rect(spans: list[dict], desc_cell_x0: float, city_cell_x1: float) -> fitz.Rect | None:
    """Get redaction rectangle for exchange rate row covering desc+city columns."""
    if not spans:
        return None

    # Find "Wisselkoers" span and the value span
    row_y0 = min(s["bbox"].y0 for s in spans)
    row_y1 = max(s["bbox"].y1 for s in spans)

    # Return rectangle covering the description and city column area
    return fitz.Rect(desc_cell_x0, row_y0 - 1, city_cell_x1, row_y1 + 1)


def parse_exchange_rate_row(spans: list[dict]) -> dict | None:
    """Parse exchange rate row and extract rate info."""
    if not spans:
        return None

    # Combine all text to parse
    full_text = " ".join(s["text"] for s in spans)

    # Try different patterns for exchange rate lines
    # Pattern 1: "1,01437 Wisselkoers USD" (rate before text)
    match = re.search(r"([\d,\.]+)\s+Wisselkoers\s+([A-Z]{3})", full_text)
    if match:
        return {
            "rate": match.group(1),
            "currency": match.group(2),
            "full_text": full_text
        }

    # Pattern 2: "Wisselkoers USD 1,0482" (rate after currency)
    match = re.search(r"Wisselkoers\s+([A-Z]{3})\s+([\d,\.]+)", full_text)
    if match:
        return {
            "currency": match.group(1),
            "rate": match.group(2),
            "full_text": full_text
        }

    # Pattern 3: "Wisselkoers USD 1 EUR = 1,0482 USD"
    match = re.search(r"Wisselkoers\s+([A-Z]{3}).*?=\s*([\d,\.]+)", full_text)
    if match:
        return {
            "currency": match.group(1),
            "rate": match.group(2),
            "full_text": full_text
        }

    return {"full_text": full_text}


def redact_pdf(
    input_path: Path,
    output_path: Path,
    config: dict,
    dry_run: bool = False
) -> dict:
    """
    Redact sensitive information from a PDF.
    Returns statistics about what was redacted.
    """
    doc = fitz.open(input_path)
    settings = get_redaction_settings(config)
    fill_color = settings["color"]
    include_exchange_rate = settings["include_exchange_rate"]

    stats = {
        "pages": len(doc),
        "transactions_found": 0,
        "transactions_redacted": 0,
        "transactions_kept": 0,
        "exchange_rates_redacted": 0,
        "redacted_items": [],
        "kept_items": []
    }

    for page_num, page in enumerate(doc):
        spans = extract_text_spans(page)
        rows = group_spans_by_row(spans)
        row_keys = list(rows.keys())

        # Track which row indices have been redacted/kept (for exchange rate detection)
        redacted_row_info = {}  # y -> parsed info
        kept_row_info = {}  # y -> (parsed info, idx in kept_items)

        for idx, y in enumerate(row_keys):
            row_spans = rows[y]

            if not is_transaction_row(row_spans):
                continue

            stats["transactions_found"] += 1
            parsed = parse_transaction_row(row_spans)

            if parsed is None:
                continue

            match, term = should_redact(parsed["description"], parsed["city"], config)

            if match:
                stats["transactions_redacted"] += 1
                redacted_row_info[y] = parsed
                stats["redacted_items"].append({
                    "page": page_num + 1,
                    "description": parsed["description"],
                    "city": parsed["city"],
                    "matched_term": term,
                    "has_exchange_rate": False
                })

                if not dry_run:
                    # Use full cell width for redaction
                    desc_rect = parsed["desc_cell"]
                    city_rect = parsed["city_cell"]

                    page.add_redact_annot(desc_rect, fill=fill_color)
                    page.add_redact_annot(city_rect, fill=fill_color)
            else:
                # Track kept (non-redacted) transactions for report
                stats["transactions_kept"] += 1
                kept_item = {
                    "page": page_num + 1,
                    "source_file": input_path.name,
                    "date1": parsed["date1"],
                    "date2": parsed["date2"],
                    "description": parsed["description"],
                    "city": parsed["city"],
                    "country": parsed["country"],
                    "amounts": parsed["amounts"],
                    "exchange_rate": None
                }
                kept_row_info[y] = (parsed, len(stats["kept_items"]))
                stats["kept_items"].append(kept_item)

        # Second pass: find exchange rate lines
        for idx, y in enumerate(row_keys):
            row_spans = rows[y]

            if not is_exchange_rate_row(row_spans):
                continue

            # Check if previous row was a transaction
            if idx > 0:
                prev_y = row_keys[idx - 1]

                # Handle redacted transactions
                if prev_y in redacted_row_info and include_exchange_rate:
                    parsed = redacted_row_info[prev_y]
                    stats["exchange_rates_redacted"] += 1

                    # Mark that this redacted item has an exchange rate
                    for item in stats["redacted_items"]:
                        if (item["description"] == parsed["description"] and
                            item["city"] == parsed["city"]):
                            item["has_exchange_rate"] = True
                            break

                    if not dry_run:
                        # Get the exchange rate redaction rect using the same column boundaries
                        exch_rect = get_exchange_rate_redact_rect(
                            row_spans,
                            parsed["desc_cell"].x0,
                            parsed["city_cell"].x1
                        )
                        if exch_rect:
                            page.add_redact_annot(exch_rect, fill=fill_color)

                # Handle kept transactions - capture exchange rate info
                elif prev_y in kept_row_info:
                    _, kept_idx = kept_row_info[prev_y]
                    exch_info = parse_exchange_rate_row(row_spans)
                    if exch_info:
                        stats["kept_items"][kept_idx]["exchange_rate"] = exch_info

        if not dry_run:
            # Apply all redactions on this page (removes underlying text)
            page.apply_redactions()

    if not dry_run:
        # Save with garbage collection and compression
        doc.save(output_path, garbage=4, deflate=True)

    doc.close()

    # Verify redaction by reading back the saved file
    if not dry_run and stats["transactions_redacted"] > 0:
        verification = verify_redaction(output_path, stats["redacted_items"])
        stats["verification"] = verification

    return stats


def create_excel_report(kept_items: list[dict], output_path: Path) -> None:
    """Create Excel report with non-redacted transactions."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # Headers
    headers = [
        "Source File",
        "Date",
        "Description",
        "City",
        "Country",
        "Foreign Amount",
        "EUR Amount",
        "Exchange Rate",
        "Currency"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    def parse_euro_number(s: str) -> float | None:
        """Convert European number format (comma as decimal) to float."""
        if not s:
            return None
        try:
            # Remove spaces, replace comma with period
            cleaned = s.strip().replace(" ", "").replace(".", "").replace(",", ".")
            return float(cleaned)
        except ValueError:
            return None

    # Data rows
    for row_idx, item in enumerate(kept_items, 2):
        amounts = item.get("amounts", [])

        # Parse amounts intelligently
        # Possible formats:
        # EUR only: ['38,70', 'Af'] or ['38,70', 'Bij']
        # Foreign: ['36,00', 'USD', '35,49', 'Af'] - foreign, currency, EUR, direction
        # Or: ['36,00', '35,49', 'Af', 'USD'] - depends on PDF structure

        foreign_amount = None
        eur_amount = None
        currency = ""

        # Find currency code (3 uppercase letters)
        currency_idx = None
        for i, val in enumerate(amounts):
            if re.match(r'^[A-Z]{3}$', val):
                currency = val
                currency_idx = i
                break

        # Find numeric amounts (contain digits and comma/period)
        numeric_amounts = []
        for i, val in enumerate(amounts):
            if re.match(r'^[\d\s,\.]+$', val.strip()):
                numeric_amounts.append((i, val.strip()))

        if currency and len(numeric_amounts) >= 2:
            # Foreign currency transaction - first numeric is foreign, second is EUR
            foreign_amount = parse_euro_number(numeric_amounts[0][1])
            eur_amount = parse_euro_number(numeric_amounts[1][1])
        elif len(numeric_amounts) >= 1:
            # EUR only transaction
            eur_amount = parse_euro_number(numeric_amounts[0][1])

        # Exchange rate info (may override currency from amounts)
        exch_rate = item.get("exchange_rate") or {}
        rate_value = parse_euro_number(exch_rate.get("rate", ""))
        if exch_rate.get("currency"):
            currency = exch_rate.get("currency", currency)

        row_data = [
            item.get("source_file", ""),
            item.get("date1", ""),
            item.get("description", ""),
            item.get("city", ""),
            item.get("country", ""),
            foreign_amount,
            eur_amount,
            rate_value,
            currency
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for row in range(1, len(kept_items) + 2):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Redact sensitive information from credit card bill PDFs"
    )
    parser.add_argument(
        "input_files",
        nargs="+",
        type=Path,
        help="PDF file(s) to process"
    )
    parser.add_argument(
        "--config", "-c",
        type=Path,
        default=Path("config.yaml"),
        help="Path to configuration YAML file (default: config.yaml)"
    )
    parser.add_argument(
        "--output", "-o",
        type=Path,
        help="Output directory (default: same as input, with _redacted suffix)"
    )
    parser.add_argument(
        "--suffix", "-s",
        type=str,
        default="_redacted",
        help="Suffix to add to output filename (default: _redacted)"
    )
    parser.add_argument(
        "--no-suffix",
        action="store_true",
        help="Don't add any suffix to output filename"
    )
    parser.add_argument(
        "--dry-run", "-n",
        action="store_true",
        help="Show what would be redacted without modifying files"
    )
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Show detailed output"
    )
    parser.add_argument(
        "--report",
        type=Path,
        help="Generate Excel report of non-redacted transactions to specified path"
    )

    args = parser.parse_args()

    # Load configuration
    config = load_config(args.config)
    exceptions = config.get("exceptions") or config.get("redact_terms") or []
    exact_match = config.get("exact_match") or []
    default_mode = config.get("redaction", {}).get("default", "none")
    print(f"Mode: default={default_mode}")
    print(f"Loaded {len(exceptions)} exceptions")
    print(f"Loaded {len(exact_match)} exact match patterns")

    if args.dry_run:
        print("\n[DRY RUN MODE - No files will be modified]\n")

    # Create output directory if specified
    if args.output:
        args.output.mkdir(parents=True, exist_ok=True)

    # Collect all kept items for report
    all_kept_items = []

    # Process each input file
    for input_path in args.input_files:
        if not input_path.exists():
            print(f"Error: File not found: {input_path}")
            continue

        if not input_path.suffix.lower() == ".pdf":
            print(f"Skipping non-PDF file: {input_path}")
            continue

        # Determine output path
        suffix = "" if args.no_suffix else args.suffix
        if args.output:
            output_path = args.output / f"{input_path.stem}{suffix}.pdf"
        else:
            output_path = input_path.parent / f"{input_path.stem}{suffix}.pdf"

        # Prevent overwriting source file
        if output_path.resolve() == input_path.resolve():
            print(f"Error: Output would overwrite source file: {input_path}")
            print(f"  Use --suffix to add a suffix, or --output to specify a different directory")
            continue

        print(f"\nProcessing: {input_path}")

        try:
            stats = redact_pdf(input_path, output_path, config, args.dry_run)

            print(f"  Pages: {stats['pages']}")
            print(f"  Transactions found: {stats['transactions_found']}")
            print(f"  Transactions redacted: {stats['transactions_redacted']}")
            if stats.get('exchange_rates_redacted', 0) > 0:
                print(f"  Exchange rates redacted: {stats['exchange_rates_redacted']}")

            if args.verbose or args.dry_run:
                for item in stats["redacted_items"]:
                    exch_note = " (+exchange rate)" if item.get("has_exchange_rate") else ""
                    print(f"    [Page {item['page']}] {item['description'][:40]}... | "
                          f"{item['city']} (matched: {item['matched_term']}){exch_note}")

            # Collect kept items for report
            all_kept_items.extend(stats.get("kept_items", []))

            if not args.dry_run:
                print(f"  Output: {output_path}")

                # Show verification results
                verification = stats.get("verification", {})
                if verification:
                    if verification.get("passed"):
                        print(f"  Verification: PASSED (checked {len(verification.get('checked_terms', []))} terms)")
                    else:
                        print(f"  Verification: FAILED!")
                        if verification.get("leaked_terms"):
                            print(f"    Leaked terms found: {verification['leaked_terms']}")
                        if verification.get("error"):
                            print(f"    Error: {verification['error']}")

        except Exception as e:
            print(f"  Error processing file: {e}")
            if args.verbose:
                import traceback
                traceback.print_exc()
            continue

    # Generate Excel report if requested
    if args.report and all_kept_items:
        try:
            create_excel_report(all_kept_items, args.report)
            print(f"\nReport: {args.report} ({len(all_kept_items)} transactions)")
        except Exception as e:
            print(f"\nError creating report: {e}")
            if args.verbose:
                import traceback
                traceback.print_exc()
    elif args.report and not all_kept_items:
        print("\nNo transactions to report (all were redacted)")

    print("\nDone.")


if __name__ == "__main__":
    main()
